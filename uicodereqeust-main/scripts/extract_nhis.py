import csv
import json
import re
import sys
import time
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


FIELDS = [
    "policy_number",
    "member_type",
    "first_name",
    "surname",
    "full_name",
    "gender",
    "dob",
    "hcp_code",
]

HYPHENS = dict.fromkeys(map(ord, "\u2010\u2011\u2012\u2013\u2014"), "-")


def normalize_line(value: str) -> str:
    return " ".join(value.replace("\u00a0", " ").translate(HYPHENS).split())


def build_record(row_match, current_hcp):
    nhia_number = row_match.group(2)
    policy_number = nhia_number.split("-")[0]
    member_type = row_match.group(3).upper().replace("  ", " ")
    name_tokens = row_match.group(4).strip().split()
    first_name = " ".join(name_tokens[:-1]).strip()
    surname = name_tokens[-1].strip()
    return {
        "policy_number": policy_number,
        "member_type": member_type,
        "first_name": first_name,
        "surname": surname,
        "full_name": f"{surname} {first_name}".strip(),
        "gender": row_match.group(5).upper(),
        "dob": row_match.group(6),
        "hcp_code": current_hcp,
    }


def parse_pdf(pdf_path: Path):
    reader = PdfReader(str(pdf_path))
    records = []
    current_hcp = ""

    row_pattern = re.compile(
        r"^\s*(\d+)\s+([0-9]+(?:-[0-9]*)*)\s+"
        r"(PRINCIPAL|SPOUSE|MEMBER|GIFSHIP|CHILD(?:\s+\d+)?|EXTRA\s+DEPENDENT(?:\s+\d+)?)\s+"
        r"(.+?)\s+([MF])\s+(\d{2}/\d{2}/\d{4})(?:\s+(\S+))?\s*$",
        re.IGNORECASE,
    )
    provider_pattern = re.compile(r"Provider Number:\s*([A-Z]{2,3}/\d{4}/P)", re.IGNORECASE)

    for page in reader.pages:
        lines = [normalize_line(raw_line) for raw_line in (page.extract_text() or "").splitlines()]
        index = 0
        while index < len(lines):
            line = lines[index]
            provider_match = provider_pattern.search(line)
            if provider_match:
                current_hcp = provider_match.group(1).upper()
                index += 1
                continue

            row_match = row_pattern.match(line)
            if not row_match and index + 1 < len(lines):
                row_match = row_pattern.match(f"{line}{lines[index + 1]}")
                if row_match:
                    index += 1

            if row_match:
                records.append(build_record(row_match, current_hcp))

            index += 1

    return records


def validate(records):
    missing = 0
    invalid_dates = 0
    seen = set()
    duplicates = 0
    hcp_counts = Counter()

    for record in records:
        if any(not str(record.get(field, "")).strip() for field in FIELDS if field != "hcp_code"):
            missing += 1
        try:
            datetime.strptime(record["dob"], "%d/%m/%Y")
        except ValueError:
            invalid_dates += 1
        key = tuple(record.get(field, "") for field in FIELDS)
        if key in seen:
            duplicates += 1
        seen.add(key)
        hcp_counts[record.get("hcp_code", "") or "Unknown"] += 1

    return {
        "total_records": len(records),
        "unique_policy_numbers": len({record["policy_number"] for record in records}),
        "duplicate_records": duplicates,
        "missing_fields": missing,
        "invalid_dates": invalid_dates,
        "hcp_summary": dict(hcp_counts.most_common(25)),
    }


def write_csv(records, output_path: Path):
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(records)


def write_xlsx(records, output_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Beneficiaries"
    sheet.append(FIELDS)

    for record in records:
        sheet.append([record.get(field, "") for field in FIELDS])

    header_fill = PatternFill("solid", fgColor="0F766E")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [16, 18, 22, 22, 34, 10, 14, 14]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    workbook.save(output_path)


def main():
    pdf_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("nhis_new.pdf")
    output_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("outputs") / "nhis"
    output_dir.mkdir(parents=True, exist_ok=True)

    started = time.perf_counter()
    records = parse_pdf(pdf_path)
    summary = validate(records)

    csv_path = output_dir / "extracted_beneficiaries.csv"
    xlsx_path = output_dir / "extracted_beneficiaries.xlsx"
    json_path = output_dir / "extraction_summary.json"

    write_csv(records, csv_path)
    write_xlsx(records, xlsx_path)
    summary["processing_seconds"] = round(time.perf_counter() - started, 2)
    json_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    print(json.dumps({"csv": str(csv_path), "xlsx": str(xlsx_path), **summary}, indent=2))


if __name__ == "__main__":
    main()
